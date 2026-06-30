"""Extract field dictionaries from the Excel codebooks bundled inside the survey
ZIPs (raw_pdfs_zip/). Surveys (Society12 TLSA, Society12-1 SEBAS, Society10 NHIS,
Health48 TBCS, …) ship their variable dictionary as .xls/.xlsx 編碼簿 / 譯碼本 /
_CB / 欄位說明 files rather than as a PDF field table.

Handles four common layouts via keyword-mapped header detection:
  1. Field | Len. | Type | Description | Note               (TLSA)
  2. 序號 | 英文欄位名稱 | 型態 | 欄位說明 | 欄位代碼描述 | 補充說明   (SEBAS)
  3. 章節 | 題號 | 題目 | 題型 | 選項                          (NHIS questionnaire)
  4. 變數名稱 | 欄寬 | 代碼說明 | 選項說明 | 備註                  (TBCS)

Writes extracted/<code>/<version_id>.json for each codebook that yields fields,
and appends manifest entries so prepare-data.js surfaces them.
"""
from __future__ import annotations
import io
import json
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import xlrd  # for legacy .xls
except ImportError:
    xlrd = None

ROOT = Path(__file__).resolve().parent.parent
ZIP_DIR = ROOT / "raw_pdfs_zip"
EXTRACTED = ROOT / "extracted"
MANIFEST = ROOT / "manifest.json"

# Column-role keyword aliases (matched case-insensitively, substring).
ROLE_KEYWORDS = {
    "name_en":  ["英文欄位名稱", "變數名稱", "變項名稱", "欄位名稱", "field", "variable", "var name", "題號"],
    "name_zh":  ["中文欄位名稱", "變項標籤", "中文名稱"],
    "desc":     ["欄位說明", "代碼說明", "題目", "description", "說明", "label", "變項說明"],
    "type":     ["型態", "資料型態", "type"],
    "length":   ["欄寬", "欄位長度", "長度", "len", "len.", "width"],
    "values":   ["欄位代碼描述", "代碼描述", "選項說明", "選項", "代碼", "value", "note", "備註", "數值標籤"],
}

# Skip clearly non-codebook spreadsheets (attachments / lookup tables / file lists).
SKIP_NAME_PATTERNS = [
    "資料內容", "親屬", "關係代號", "鄉鎮", "縣市", "代號表", "sudaan",
    "id對照", "id 對照", "對照檔", "icd", "職業分類", "過錄代號",
]

# Only treat as a codebook if the filename hints at one (keeps version count sane).
KEEP_NAME_PATTERNS = [
    "codebook", "編碼簿", "譯碼本", "譯碼簿", "_cb", "欄位說明", "變項", "code book",
]


def norm(s) -> str:
    return re.sub(r"\s+", "", str(s or "")).strip().lower()


def detect_header(rows: list[list]) -> tuple[int, dict] | None:
    """Find the header row in the first ~12 rows; return (row_idx, {role: col_idx})."""
    for i, row in enumerate(rows[:12]):
        mapping: dict[str, int] = {}
        for c, cell in enumerate(row):
            t = norm(cell)
            if not t:
                continue
            for role, kws in ROLE_KEYWORDS.items():
                if role in mapping:
                    continue
                if any(k in t for k in kws):
                    mapping[role] = c
                    break
        # Need at least a name/variable column and one of desc/type/values
        has_name = "name_en" in mapping or "name_zh" in mapping
        has_meat = any(r in mapping for r in ("desc", "type", "values", "length"))
        if has_name and has_meat:
            return i, mapping
    return None


def rows_from_xlsx(path: Path) -> list[tuple[str, list[list]]]:
    out = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sh in wb.sheetnames:
        ws = wb[sh]
        rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
        out.append((sh, rows))
    wb.close()
    return out


def rows_from_xls(path: Path) -> list[tuple[str, list[list]]]:
    if xlrd is None:
        return []
    out = []
    wb = xlrd.open_workbook(str(path))
    for sh in wb.sheet_names():
        ws = wb.sheet_by_name(sh)
        rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
        out.append((sh, rows))
    return out


def parse_sheet(rows: list[list], seq_start: int = 1) -> list[dict]:
    det = detect_header(rows)
    if not det:
        return []
    hidx, m = det
    fields = []
    seq = seq_start
    for row in rows[hidx + 1:]:
        def get(role):
            ci = m.get(role)
            if ci is None or ci >= len(row):
                return ""
            return str(row[ci]).strip() if row[ci] is not None else ""
        name_en = get("name_en")
        name_zh = get("name_zh")
        desc = get("desc")
        # A real field row has a variable name (en) or a Chinese label.
        primary = name_en or name_zh or desc
        if not primary:
            continue
        # Skip section-divider rows (only a Chinese heading, no variable token)
        if not name_en and not re.search(r"[A-Za-z0-9_]", primary):
            continue
        length = get("length").replace(".0", "")
        typ = get("type")
        values = get("values")
        f = {
            "seq": seq,
            "name_zh": name_zh or "",
            "name_en": name_en or "",
            "type": typ,
            "length": int(length) if length.isdigit() else (length or ""),
            "description_zh": desc or "",
        }
        # Fold value labels into description if present (kept inline; codebook UI optional)
        if values:
            f["value_labels"] = values
        f["available_notes"] = []
        fields.append(f)
        seq += 1
    return fields


def parse_excel(path: Path) -> list[dict]:
    try:
        sheets = rows_from_xlsx(path) if path.suffix.lower() in (".xlsx", ".xlsm") \
            else rows_from_xls(path)
    except Exception as e:
        print(f"    ! read error {path.name}: {e}", file=sys.stderr)
        return []
    all_fields = []
    seq = 1
    for sh_name, rows in sheets:
        fs = parse_sheet(rows, seq_start=seq)
        if fs:
            # Tag the section/sheet so multi-sheet questionnaires stay navigable
            for f in fs:
                if sh_name and sh_name not in ("工作表1", "Sheet1"):
                    f["section"] = sh_name
            all_fields.extend(fs)
            seq += len(fs)
    return all_fields


def safe(name: str) -> str:
    return re.sub(r"[^\w\-]", "_", name, flags=re.UNICODE)[:80]


def wave_from_name(stem: str) -> str:
    """Return an AD-normalised wave tag so versions sort chronologically.
    ROC years (民國NN / 2-3 digit) are converted to AD (NN + 1911)."""
    # Explicit AD year first
    m = re.search(r"(20\d{2}|19\d{2})", stem)
    if m:
        return f"AD{m.group(1)}"
    # ROC year: 民國NN年 / NN年
    m = re.search(r"民國\s*(\d{2,3})\s*年", stem) or re.search(r"(\d{2,3})\s*年", stem)
    if m:
        return f"AD{int(m.group(1)) + 1911}"
    # Age-based waves (TBCS N歲) — no calendar year; keep as-is, sort late
    m = re.search(r"(\d+)\s*歲", stem)
    if m:
        return f"age{int(m.group(1)):02d}"
    return safe(stem)[:24]


def main():
    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    versions = manifest["versions"]
    have = {(v["code"], v["version_id"]) for v in versions}

    # Map each zip dir to its (code, parent version_id)
    zip_versions = [v for v in versions if v.get("file_type") == "zip"]
    added = 0
    scanned = 0

    excel_paths = []
    for ext in ("*.xlsx", "*.xls", "*.xlsm"):
        excel_paths += ZIP_DIR.rglob(ext)

    # Dedupe: the legacy and post-20250624 ZIPs bundle identical Excel files.
    # Keep one per (code, filename), preferring the post-20250624 parent.
    chosen: dict[tuple[str, str], Path] = {}
    for xp in excel_paths:
        stem = xp.stem
        low = stem.lower()
        if low.startswith("~$"):
            continue
        if any(p in low for p in SKIP_NAME_PATTERNS):
            continue
        if not any(p in low for p in KEEP_NAME_PATTERNS):
            continue
        rel = xp.relative_to(ZIP_DIR)
        top = rel.parts[0]
        mcode = re.match(r"((?:Health|Society|Welfare)[\w\-]*?)_", top)
        if not mcode:
            continue
        code = mcode.group(1)
        key = (code, stem)
        prev = chosen.get(key)
        if prev is None or ("post-20250624" in top and "post-20250624" not in str(prev)):
            chosen[key] = xp

    for (code, stem), xp in sorted(chosen.items()):
        rel = xp.relative_to(ZIP_DIR)
        top = rel.parts[0]
        # Always present the dedup'd codebook under the file's most-current revision.
        parent_vid = "post-20250624" if "post-20250624" in top else "legacy"
        low = stem.lower()

        scanned += 1
        fields = parse_excel(xp)
        if len(fields) < 3:
            continue

        wave = wave_from_name(stem)
        version_id = f"{parent_vid}__xls_{wave}_{safe(stem)[:20]}"
        if (code, version_id) in have:
            continue

        data = {
            "source_pdf": xp.name,
            "code": code,
            "category": next((v["category"] for v in versions if v["code"] == code), "Unknown"),
            "code_short": "",
            "name_zh": stem,
            "name_en": "",
            "version_id": version_id,
            "version_label": f"Excel codebook · {stem}",
            "record_count_raw": "",
            "field_count_declared": None,
            "frequency": "",
            "attribute": "Excel codebook (survey)",
            "update_history": [],
            "data_description_zh": "",
            "notes_zh": "",
            "primary_keys_raw": "",
            "fields": fields,
            "codebooks": {},
            "qa": {"errors": [], "warnings": []},
            "source_excel": str(xp.relative_to(ROOT)).replace("\\", "/"),
        }
        out_dir = EXTRACTED / code
        out_dir.mkdir(parents=True, exist_ok=True)
        (out_dir / f"{safe(version_id)}.json").write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

        versions.append({
            "code": code,
            "category": data["category"],
            "name_zh": stem,
            "version_id": version_id,
            "version_label": data["version_label"],
            "detail_url": next((v["detail_url"] for v in versions if v["code"] == code), ""),
            "pdf_url": "",
            "file_type": "xls",
            "doc_first_published": "",
            "doc_last_updated": "",
            "sha256": "",
            "local_filename": data["source_excel"],
            "fetched_at": "from-zip-excel",
            "note": f"excel codebook from {top}",
        })
        have.add((code, version_id))
        added += 1
        print(f"  [OK ] {code:12s} {wave:10s} {len(fields):>4} fields  {stem[:40]}")

    manifest["versions"] = versions
    manifest["count"] = len(versions)
    MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nScanned {scanned} candidate Excel files; added {added} codebook versions.")


if __name__ == "__main__":
    main()
