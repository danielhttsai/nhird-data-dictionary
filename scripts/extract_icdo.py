"""Build ICD-O topography (T-code) and morphology (M-code) reference codebooks
from the TWCR Excel tables, and attach them to the cancer-registry fields that
reference them (原發部位 → topography, 組織型態 → morphology).
"""
import io, json, re, sys
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(__file__).resolve().parent.parent
TW = ROOT / "raw_pdfs_twcr"
EXTRACTED = ROOT / "extracted"


def load_topography():
    wb = openpyxl.load_workbook(TW / "ICDO_topography_Tcode.xlsx", read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    entries = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        ft, casite, name = (row + (None, None, None))[:3]
        code = str(casite or ft or "").strip()
        label = re.sub(r"\s+", " ", str(name or "").strip())
        if code and label:
            entries.append({"code": code, "label_zh": label})
    wb.close()
    return entries


def load_morphology():
    wb = openpyxl.load_workbook(TW / "ICDO_morphology_Mcode.xlsx", read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    entries = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        mcode, name = (row + (None, None))[:2]
        code = str(mcode or "").strip()
        label = re.sub(r"\s+", " ", str(name or "").strip())
        if code and label:
            entries.append({"code": code, "label_zh": label})
    wb.close()
    return entries


def main():
    topo = load_topography()
    morph = load_morphology()
    print(f"ICD-O topography: {len(topo)} codes | morphology: {len(morph)} codes")

    icdo = {
        "topography": {"field_zh": "原發部位 (ICD-O Topography)", "field_en": "ICD-O-3 Topography (T-code)",
                       "source": "TWCR ICD-O T-code 對照表", "entries": topo},
        "morphology": {"field_zh": "組織型態 (ICD-O Morphology)", "field_en": "ICD-O-3 Morphology (M-code)",
                       "source": "TWCR ICD-O M-code 對照表", "entries": morph},
    }

    # attach to cancer-registry files: topography → 原發部位, morphology → 組織型態/形態
    targets = ["Health14/twcr_manual_LF129_2025", "Health14/twcr_manual_LF115_2018",
               "Health15/twcr_manual_SF50_2025", "Health15/twcr_manual_SF45_2018"]
    # also MOHW released-file cancer registry versions
    import glob
    for code in ("Health14", "Health15", "Health16"):
        for p in glob.glob(str(EXTRACTED / code / "*.json")):
            if "diff_" in p or "twcr" in p:
                continue
            targets.append(str(Path(p).relative_to(EXTRACTED)).replace("\\", "/").replace(".json", ""))

    for t in targets:
        p = EXTRACTED / f"{t}.json"
        if not p.exists():
            continue
        d = json.loads(p.read_text(encoding="utf-8"))
        cbs = d.get("codebooks", {})
        n = 0
        for f in d.get("fields", []):
            nz = f.get("name_zh", "")
            key = f.get("name_en") or f.get("seq")
            if "原發部位" in nz and "手術" not in nz and "未" not in nz:
                cbs[key] = dict(icdo["topography"]); n += 1
            elif ("組織型態" in nz) or ("病理形態" in nz) or (nz.strip() == "形態"):
                cbs[key] = dict(icdo["morphology"]); n += 1
        if n:
            d["codebooks"] = cbs
            p.write_text(json.dumps(d, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"  {t}: attached ICD-O to {n} field(s)")


if __name__ == "__main__":
    main()
